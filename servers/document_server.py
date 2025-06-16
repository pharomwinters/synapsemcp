#
# This file is part of Synapse.
#
# Synapse is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Synapse is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Synapse.  If not, see <https://www.gnu.org/licenses/>.
#
"""
Document Server - Document management for Synapse.

This server handles document storage, retrieval, and management operations
for the Synapse platform. Supports various document formats including
MS Office, LibreOffice, PDF, and text files.
"""

from fastmcp import FastMCP
from typing import Dict, Any, List, Optional
import os
import hashlib
import mimetypes
from pathlib import Path
from datetime import datetime
import shutil

# Document processing imports
try:
    import PyPDF2
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from odf import text, teletype
    from odf.opendocument import load as odf_load
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False

try:
    import win32com.client
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False

# Local imports
from config import get_config
from utils import get_database

# Create Document Server instance
document_server = FastMCP("Document Synapse")

# Supported file extensions
SUPPORTED_EXTENSIONS = {
    # Microsoft Office
    '.docx': 'Microsoft Word Document',
    '.doc': 'Microsoft Word Document (Legacy)',
    '.xlsx': 'Microsoft Excel Spreadsheet', 
    '.xls': 'Microsoft Excel Spreadsheet (Legacy)',
    '.pptx': 'Microsoft PowerPoint Presentation',
    '.ppt': 'Microsoft PowerPoint Presentation (Legacy)',
    
    # LibreOffice/OpenOffice
    '.odt': 'OpenDocument Text',
    '.ods': 'OpenDocument Spreadsheet',
    '.odp': 'OpenDocument Presentation',
    '.odg': 'OpenDocument Graphics',
    '.odf': 'OpenDocument Formula',
    
    # PDF
    '.pdf': 'Portable Document Format',
    
    # Text formats
    '.txt': 'Plain Text',
    '.rtf': 'Rich Text Format',
    '.csv': 'Comma Separated Values',
    '.xml': 'XML Document',
    '.html': 'HTML Document',
    '.htm': 'HTML Document',
    '.md': 'Markdown Document',
    '.markdown': 'Markdown Document',
}

def get_file_hash(file_path: str) -> str:
    """Generate SHA-256 hash of file content."""
    hasher = hashlib.sha256()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hasher.update(chunk)
    return hasher.hexdigest()

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file."""
    if not PDF_AVAILABLE:
        return "PDF processing not available. Install PyPDF2 and pdfplumber."
    
    text = ""
    try:
        # Try pdfplumber first (better text extraction)
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception:
        # Fallback to PyPDF2
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            return f"Error extracting PDF text: {str(e)}"
    
    return text.strip()

def extract_text_from_docx(file_path: str) -> str:
    """Extract text from DOCX file."""
    if not DOCX_AVAILABLE:
        return "DOCX processing not available. Install python-docx."
    
    try:
        doc = DocxDocument(file_path)
        text = []
        for paragraph in doc.paragraphs:
            text.append(paragraph.text)
        return "\n".join(text)
    except Exception as e:
        return f"Error extracting DOCX text: {str(e)}"

def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from XLSX file."""
    if not XLSX_AVAILABLE:
        return "XLSX processing not available. Install openpyxl."
    
    try:
        workbook = load_workbook(file_path, data_only=True)
        text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text.append(f"Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text.append("\t".join(row_text))
        return "\n".join(text)
    except Exception as e:
        return f"Error extracting XLSX text: {str(e)}"

def extract_text_from_odt(file_path: str) -> str:
    """Extract text from ODT file."""
    if not ODF_AVAILABLE:
        return "ODT processing not available. Install odfpy."
    
    try:
        doc = odf_load(file_path)
        text_elements = doc.getElementsByType(text.P)
        text_content = []
        for element in text_elements:
            text_content.append(teletype.extractText(element))
        return "\n".join(text_content)
    except Exception as e:
        return f"Error extracting ODT text: {str(e)}"

def extract_text_from_file(file_path: str, file_extension: str) -> str:
    """Extract text content from various file formats."""
    try:
        if file_extension == '.pdf':
            return extract_text_from_pdf(file_path)
        elif file_extension == '.docx':
            return extract_text_from_docx(file_path)
        elif file_extension == '.xlsx':
            return extract_text_from_xlsx(file_path)
        elif file_extension == '.odt':
            return extract_text_from_odt(file_path)
        elif file_extension in ['.txt', '.md', '.markdown', '.csv', '.xml', '.html', '.htm']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        elif file_extension == '.rtf':
            # Basic RTF text extraction (strips RTF formatting)
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                # Simple RTF text extraction - remove RTF control words
                import re
                text = re.sub(r'\\[a-z]+\d*\s?', '', content)
                text = re.sub(r'[{}]', '', text)
                return text.strip()
        else:
            return f"Text extraction not supported for {file_extension} files."
    except Exception as e:
        return f"Error reading file: {str(e)}"

@document_server.tool()
def store_document(file_path: str, document_name: Optional[str] = None, tags: Optional[List[str]] = None) -> Dict[str, Any]:
    """Store a document in the Synapse document management system.
    
    Args:
        file_path: Full path to the document file to store
        document_name: Optional custom name for the document (defaults to filename)
        tags: Optional list of tags to associate with the document
        
    Returns:
        Dictionary with storage result and document metadata
    """
    try:
        # Validate file exists
        if not os.path.exists(file_path):
            return {"success": False, "error": f"File not found: {file_path}"}
        
        # Get file info
        file_path_obj = Path(file_path)
        original_filename = file_path_obj.name
        file_extension = file_path_obj.suffix.lower()
        file_size = os.path.getsize(file_path)
        
        # Check if file type is supported
        if file_extension not in SUPPORTED_EXTENSIONS:
            return {
                "success": False, 
                "error": f"Unsupported file type: {file_extension}",
                "supported_types": list(SUPPORTED_EXTENSIONS.keys())
            }
        
        # Generate document name
        if not document_name:
            document_name = file_path_obj.stem
        
        # Generate file hash for deduplication
        file_hash = get_file_hash(file_path)
        
        # Extract text content
        extracted_text = extract_text_from_file(file_path, file_extension)
        
        # Get MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        
        # Create documents directory if it doesn't exist
        documents_dir = Path(get_config("documents_dir", "documents"))
        documents_dir.mkdir(exist_ok=True)
        
        # Copy file to documents directory with hash-based naming
        stored_filename = f"{file_hash}{file_extension}"
        stored_path = documents_dir / stored_filename
        
        # Copy file if not already stored
        if not stored_path.exists():
            shutil.copy2(file_path, stored_path)
        
        # Prepare document metadata
        document_metadata = {
            "document_name": document_name,
            "original_filename": original_filename,
            "file_extension": file_extension,
            "file_type": SUPPORTED_EXTENSIONS[file_extension],
            "mime_type": mime_type,
            "file_size": file_size,
            "file_hash": file_hash,
            "stored_filename": stored_filename,
            "stored_path": str(stored_path),
            "original_path": file_path,
            "extracted_text": extracted_text,
            "tags": tags or [],
            "created_at": datetime.now().isoformat(),
            "modified_at": datetime.now().isoformat()
        }
        
        # Store in database
        db = get_database()
        if hasattr(db, 'store_document'):
            db.store_document(document_name, document_metadata)
        else:
            # Fallback: store as memory with special prefix
            import json
            db.store_memory(f"document:{document_name}", json.dumps(document_metadata, indent=2))
        
        return {
            "success": True,
            "document_name": document_name,
            "file_hash": file_hash,
            "file_type": SUPPORTED_EXTENSIONS[file_extension],
            "file_size": file_size,
            "stored_path": str(stored_path),
            "text_preview": extracted_text[:500] + "..." if len(extracted_text) > 500 else extracted_text
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to store document: {str(e)}"}

@document_server.tool()
def get_document(document_name: str) -> Dict[str, Any]:
    """Retrieve document metadata and content.
    
    Args:
        document_name: Name of the document to retrieve
        
    Returns:
        Dictionary with document metadata and content
    """
    try:
        db = get_database()
        
        if hasattr(db, 'get_document'):
            document_data = db.get_document(document_name)
        else:
            # Fallback: get from memory
            document_json = db.get_memory(f"document:{document_name}")
            if document_json:
                import json
                document_data = json.loads(document_json)
            else:
                document_data = None
        
        if not document_data:
            return {"success": False, "error": f"Document not found: {document_name}"}
        
        # Check if stored file still exists
        stored_path = document_data.get("stored_path")
        if stored_path and not os.path.exists(stored_path):
            return {"success": False, "error": f"Stored file missing: {stored_path}"}
        
        return {
            "success": True,
            "document": document_data
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to retrieve document: {str(e)}"}

@document_server.tool()
def list_documents(tag_filter: Optional[str] = None) -> Dict[str, Any]:
    """List all stored documents.
    
    Args:
        tag_filter: Optional tag to filter documents by
        
    Returns:
        Dictionary with list of documents and their basic metadata
    """
    try:
        db = get_database()
        documents = []
        
        if hasattr(db, 'list_documents'):
            document_list = db.list_documents()
        else:
            # Fallback: search memories with document prefix
            memories = db.list_memories()
            document_list = []
            for memory_name in memories:
                if memory_name.startswith("document:"):
                    doc_name = memory_name[9:]  # Remove "document:" prefix
                    document_list.append(doc_name)
        
        for doc_name in document_list:
            doc_result = get_document(doc_name)
            if doc_result["success"]:
                doc_data = doc_result["document"]
                
                # Apply tag filter if specified
                if tag_filter:
                    doc_tags = doc_data.get("tags", [])
                    if tag_filter not in doc_tags:
                        continue
                
                documents.append({
                    "document_name": doc_data.get("document_name"),
                    "original_filename": doc_data.get("original_filename"),
                    "file_type": doc_data.get("file_type"),
                    "file_size": doc_data.get("file_size"),
                    "tags": doc_data.get("tags", []),
                    "created_at": doc_data.get("created_at"),
                    "modified_at": doc_data.get("modified_at")
                })
        
        return {
            "success": True,
            "documents": documents,
            "count": len(documents)
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to list documents: {str(e)}"}

@document_server.tool()
def search_documents(query: str, search_content: bool = True) -> Dict[str, Any]:
    """Search documents by name, tags, or content.
    
    Args:
        query: Search query string
        search_content: Whether to search document content (extracted text)
        
    Returns:
        Dictionary with matching documents
    """
    try:
        db = get_database()
        matching_documents = []
        
        # Get all documents
        docs_result = list_documents()
        if not docs_result["success"]:
            return docs_result
        
        query_lower = query.lower()
        
        for doc_summary in docs_result["documents"]:
            doc_name = doc_summary["document_name"]
            doc_result = get_document(doc_name)
            
            if not doc_result["success"]:
                continue
                
            doc_data = doc_result["document"]
            match_score = 0
            match_reasons = []
            
            # Search document name
            if query_lower in doc_data.get("document_name", "").lower():
                match_score += 10
                match_reasons.append("name")
            
            # Search original filename
            if query_lower in doc_data.get("original_filename", "").lower():
                match_score += 8
                match_reasons.append("filename")
            
            # Search tags
            for tag in doc_data.get("tags", []):
                if query_lower in tag.lower():
                    match_score += 5
                    match_reasons.append("tag")
            
            # Search content if requested
            if search_content:
                extracted_text = doc_data.get("extracted_text", "")
                if query_lower in extracted_text.lower():
                    match_score += 3
                    match_reasons.append("content")
            
            if match_score > 0:
                matching_documents.append({
                    "document_name": doc_data.get("document_name"),
                    "original_filename": doc_data.get("original_filename"),
                    "file_type": doc_data.get("file_type"),
                    "tags": doc_data.get("tags", []),
                    "match_score": match_score,
                    "match_reasons": match_reasons,
                    "content_preview": (doc_data.get("extracted_text", "")[:200] + "...") 
                                     if len(doc_data.get("extracted_text", "")) > 200 else doc_data.get("extracted_text", "")
                })
        
        # Sort by match score
        matching_documents.sort(key=lambda x: x["match_score"], reverse=True)
        
        return {
            "success": True,
            "query": query,
            "matches": matching_documents,
            "count": len(matching_documents)
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to search documents: {str(e)}"}

@document_server.tool()
def delete_document(document_name: str, delete_file: bool = False) -> Dict[str, Any]:
    """Delete a document from the system.
    
    Args:
        document_name: Name of the document to delete
        delete_file: Whether to also delete the stored file
        
    Returns:
        Dictionary with deletion result
    """
    try:
        # Get document info first
        doc_result = get_document(document_name)
        if not doc_result["success"]:
            return doc_result
        
        doc_data = doc_result["document"]
        stored_path = doc_data.get("stored_path")
        
        # Delete from database
        db = get_database()
        if hasattr(db, 'delete_document'):
            db.delete_document(document_name)
        else:
            # Fallback: delete from memory
            db.delete_memory(f"document:{document_name}")
        
        # Delete stored file if requested
        if delete_file and stored_path and os.path.exists(stored_path):
            os.remove(stored_path)
            file_deleted = True
        else:
            file_deleted = False
        
        return {
            "success": True,
            "document_name": document_name,
            "file_deleted": file_deleted
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to delete document: {str(e)}"}

@document_server.tool()
def get_supported_formats() -> Dict[str, Any]:
    """Get list of supported document formats and their availability.
    
    Returns:
        Dictionary with supported formats and processing capabilities
    """
    capabilities = {
        "pdf": PDF_AVAILABLE,
        "docx": DOCX_AVAILABLE,
        "xlsx": XLSX_AVAILABLE,
        "odt": ODF_AVAILABLE,
        "win32_office": WIN32_AVAILABLE
    }
    
    return {
        "success": True,
        "supported_extensions": SUPPORTED_EXTENSIONS,
        "processing_capabilities": capabilities,
        "missing_dependencies": [
            f"{fmt}: pip install {pkg}" for fmt, pkg in [
                ("pdf", "PyPDF2 pdfplumber"),
                ("docx", "python-docx"),
                ("xlsx", "openpyxl"),
                ("odt", "odfpy"),
                ("win32_office", "pywin32")
            ] if not capabilities.get(fmt.split('_')[0], True)
        ]
    }

@document_server.tool()
def add_document_tags(document_name: str, tags: List[str]) -> Dict[str, Any]:
    """Add tags to an existing document.
    
    Args:
        document_name: Name of the document
        tags: List of tags to add
        
    Returns:
        Dictionary with update result
    """
    try:
        # Get current document
        doc_result = get_document(document_name)
        if not doc_result["success"]:
            return doc_result
        
        doc_data = doc_result["document"]
        current_tags = set(doc_data.get("tags", []))
        new_tags = set(tags)
        
        # Add new tags
        updated_tags = list(current_tags.union(new_tags))
        doc_data["tags"] = updated_tags
        doc_data["modified_at"] = datetime.now().isoformat()
        
        # Update in database
        db = get_database()
        if hasattr(db, 'store_document'):
            db.store_document(document_name, doc_data)
        else:
            # Fallback: update in memory
            import json
            db.store_memory(f"document:{document_name}", json.dumps(doc_data, indent=2))
        
        return {
            "success": True,
            "document_name": document_name,
            "added_tags": list(new_tags - current_tags),
            "all_tags": updated_tags
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to add tags: {str(e)}"}

# Export the server instance
__all__ = ['document_server']

if __name__ == "__main__":
    # Run the document server standalone for testing
    document_server.run()